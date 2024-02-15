import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from math import sqrt

wb = openpyxl.load_workbook('test.xlsx')            # Loads spreadsheet

def write_ans(x, y, i):
    ws.cell(row = i, column = 1, value = x)
    ws.cell(row = i, column = 2, value = y)

def write_err(i):
    ws.cell(row = i, column = 1, value = 'This equation has complex roots')

def calculate_roots(a, b, c, i):
    if a == 0:                                     # Verify quadratic equation
        ws.cell(row = i, column = 1, value = 'Not a quadratic equation. \'a\' cannot be zero')
    else:   
        discriminant = (b)**2 - 4 * a * c           # Calculate discriminant
        if discriminant >= 0:                       # Calculates only real roots
            x_one = (-b + sqrt(discriminant)) / (2 * a)
            x_two = (-b - sqrt(discriminant)) / (2 * a)      
            write_ans(x_one, x_two, i)
        else:                                      # Complex roots error handling
            write_err(i)

def create_sheet():
    wb.create_sheet('Output')
    global ws
    ws = wb['Output']
    ws['A1'] = "x1"                                 # Header at A1
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws['B1'] = "x2"                                 # Header at B2
    ws["B1"].font = Font(bold=True)
    ws["B1"].alignment = Alignment(horizontal="center")

def clear_sheet():
    wb.remove(wb['Output'])

def read():
    clear_sheet()                                   # Resets previous calculations
    create_sheet()                                  # Creates a clean sheet
    df = pd.read_excel('test.xlsx', sheet_name = 'Sheet1')          # Reads inputs
    for i, row in df.iterrows():                    # Fetches values from each cell
        l = row.to_list()
        a, b, c = l[0], l[1], l[2]
        calculate_roots(a, b, c, i+2)               # Calculates for each row offset to the begin from the second
    wb.save('test.xlsx')
read()
